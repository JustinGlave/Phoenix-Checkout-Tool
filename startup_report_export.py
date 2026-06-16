"""
startup_report_export.py — export a job's checkout records to the Phoenix
Startup Report workbook.

Design (per CHECKOUT_STARTUP_REPORT_FEATURE_AUDIT.md, decisions approved 2026-06-16):

- The template is loaded from embedded bytes (startup_report_template.py) so the
  feature ships safely through the exe-only auto-updater. Nothing here touches the
  updater, build, or installer.
- We write PLAIN VALUES only, to the "Startup Report" tab only. We never write to
  the "Cover" tab (it auto-fills via formulas) and never write column A (the
  template auto-numbers it with =IF(B{r}="","",ROW()-14)).
- All styling, dropdowns, conditional formatting, merges, and print setup belong to
  the template and are preserved by openpyxl's load -> write -> save round-trip.

This module is intentionally Qt-free so it can be unit-tested headlessly. The Qt
metadata dialog and menu wiring live in checkout_tool_gui.py.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Iterable, Optional

import openpyxl
from openpyxl.styles import Alignment

from startup_report_template import template_stream

# ── Template layout constants (audited against the v2.0 template) ───────────────
SHEET_NAME      = "Startup Report"
FIRST_DATA_ROW  = 15
LAST_DATA_ROW   = 54
MAX_VALVES      = LAST_DATA_ROW - FIRST_DATA_ROW + 1   # 40 preformatted rows

# Metadata cells (single cells / merged top-left cells on the Startup Report tab).
META_CELLS = {
    "project":           "C3",
    "site_name":         "C4",
    "building":          "C5",
    "floor":             "C6",
    "technician":        "C7",
    "ats_job_number":    "C8",
    "date_of_checkout":  "C9",
    "product_lines":     "C10",
    "description":       "C11",
    "executive_summary": "G4",
}

# Valve row columns B..K (column A is auto-numbered). The hidden detailed-reading
# columns L..AE are never written by the app (they stay blank) — including the
# operator-flagged Mag Diff Pressure, Primary Air from AHU, Htg/Clg Offset, Exhaust
# Damper Pos., Verify Schedule, Air Differential, and Linkage Cotter Pins.
# F (Location / Room) is written from record.location_room; J (Face Velocity FPM)
# is intentionally blank for v1.

# ── Mapping tables (approved) ───────────────────────────────────────────────────
_CSCP_TYPES = {"CSCP Fume Hood", "PBC Room"}
_CELERIS_TYPES = {
    "Fume Hood", "GEX", "MAV", "Snorkel",
    "Canopy", "Draw Down Bench", "Gas Cabinet",
}

# Checkout valve_type -> Startup Report "Valve Type" dropdown value.
# Allowed dropdown values: Supply / General Exhaust / Fume Hood / PBC.
VALVE_TYPE_MAP = {
    "Fume Hood":       "Fume Hood",
    "CSCP Fume Hood":  "Fume Hood",
    "PBC Room":        "PBC",
    "MAV":             "Supply",
    "GEX":             "General Exhaust",
    "Snorkel":         "General Exhaust",
    "Canopy":          "General Exhaust",
    "Draw Down Bench": "General Exhaust",
    "Gas Cabinet":     "General Exhaust",
}

# Checkout pass_fail -> Startup Report "Pass/Fail" dropdown value (uppercase).
PASS_FAIL_MAP = {"Pass": "PASS", "Fail": "FAIL"}


class TooManyValvesError(Exception):
    """Raised when a job exceeds the v1 cap of MAX_VALVES (40) valves."""


@dataclass
class StartupReportMeta:
    """Job-level metadata for the Startup Report header block."""
    project:           str = ""
    site_name:         str = ""
    building:          str = ""
    floor:             str = ""
    technician:        str = ""
    ats_job_number:    str = ""
    date_of_checkout:  str = ""
    product_lines:     str = ""
    description:       str = ""
    executive_summary: str = ""


# ── Mapping helpers ─────────────────────────────────────────────────────────────

def map_product_line(valve_type: str) -> str:
    """CSCP / Celeris from the Checkout valve_type. '' if unknown."""
    if valve_type in _CSCP_TYPES:
        return "CSCP"
    if valve_type in _CELERIS_TYPES:
        return "Celeris"
    return ""


def map_valve_type(valve_type: str) -> str:
    """Startup Report 'Valve Type' value from the Checkout valve_type. '' if unknown."""
    return VALVE_TYPE_MAP.get(valve_type, "")


def map_pass_fail(pass_fail: str) -> str:
    """'PASS' / 'FAIL' from the Checkout pass_fail. '' when unset (leave blank)."""
    return PASS_FAIL_MAP.get(pass_fail, "")


def derive_product_lines(records: Iterable) -> str:
    """Summarize the product line(s) present across a job's records.

    'CSCP', 'Celeris', 'CSCP & Celeris', or '' if none are recognized.
    """
    lines = {map_product_line(getattr(r, "valve_type", "")) for r in records}
    lines.discard("")
    if lines == {"CSCP", "Celeris"}:
        return "CSCP & Celeris"
    if lines:
        return next(iter(lines))
    return ""


def generate_executive_summary(records: Iterable) -> str:
    """Generate a factual executive summary for cell G4 (editable before export).

    Reports total / pass / fail counts and lists the failed valve tags. Issues are
    never invented: if nothing failed, it says so. There is no job-level notes field
    in the data model, so per-valve notes are referenced ("See Notes column") rather
    than copied here.
    """
    records = list(records)
    total = len(records)
    if total == 0:
        return "No valves checked."
    passed = sum(1 for r in records if (getattr(r, "pass_fail", "") or "") == "Pass")
    failed = [r for r in records if (getattr(r, "pass_fail", "") or "") == "Fail"]
    nfail = len(failed)
    if nfail == 0:
        if passed == total:
            return "All checked valves passed. No issues noted."
        return f"{total} valves checked. {passed} passed, 0 failed. No issues noted."
    tags = [((getattr(r, "valve_tag", "") or "").strip() or "(unnamed valve)") for r in failed]
    plural_v = "s" if total != 1 else ""
    plural_i = "s" if nfail != 1 else ""
    return (
        f"{total} valve{plural_v} checked. {passed} passed, {nfail} failed. "
        f"Issue{plural_i} noted on {', '.join(tags)}. See Notes column for details."
    )


def prefill_meta(record, records: Iterable) -> StartupReportMeta:
    """Build a StartupReportMeta pre-filled from a representative checkout record.

    Project / Technician / Date / ATS Job # / Description come from `record`;
    Product Line(s) and the Executive Summary are generated from all `records`;
    Site / Building / Floor are left blank for the operator to fill in. All fields
    remain editable in the dialog before export.
    """
    rec_list = list(records)
    summary = generate_executive_summary(rec_list)
    if record is None:
        return StartupReportMeta(
            product_lines=derive_product_lines(rec_list),
            executive_summary=summary,
        )
    return StartupReportMeta(
        project=getattr(record, "project", "") or "",
        technician=getattr(record, "technician", "") or "",
        date_of_checkout=getattr(record, "date", "") or "",
        ats_job_number=getattr(record, "ats_job_number", "") or "",
        description=getattr(record, "description", "") or "",
        product_lines=derive_product_lines(rec_list),
        executive_summary=summary,
    )


# ── Cell-write helpers ──────────────────────────────────────────────────────────

def _set(ws, ref: str, value) -> None:
    """Write a value only if it is meaningful; leave the template cell untouched otherwise."""
    if value is None:
        return
    if isinstance(value, str) and value.strip() == "":
        return
    ws[ref] = value


def _num_or_text(value) -> Optional[object]:
    """Coerce a setpoint string to int/float when numeric; keep non-empty text; None if blank."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    neg = s[1:] if s.startswith("-") else s
    if neg.isdigit():
        return int(s)
    try:
        return float(s)
    except ValueError:
        return s


def _enable_notes_wrap(ws) -> None:
    """Force wrap_text on the Notes column (K) data rows, preserving other alignment.

    Only alignment is touched; fonts/fills/borders/number_format are untouched.
    """
    for r in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        cell = ws[f"K{r}"]
        al = cell.alignment
        cell.alignment = Alignment(
            horizontal=al.horizontal,
            vertical=al.vertical,
            text_rotation=al.text_rotation,
            wrap_text=True,
            shrink_to_fit=al.shrink_to_fit,
            indent=al.indent,
        )


# ── Public export entry point ───────────────────────────────────────────────────

def export_startup_report(meta: StartupReportMeta, records, output_path: str) -> None:
    """
    Populate the embedded Startup Report template with job metadata + valve rows
    and save to output_path (.xlsx).

    Raises TooManyValvesError if records exceeds the v1 cap (MAX_VALVES). Never
    writes the Cover tab or column A. Writes plain values only.
    """
    records = list(records)
    if len(records) > MAX_VALVES:
        raise TooManyValvesError(
            f"{len(records)} valves exceed the v1 template capacity of {MAX_VALVES}."
        )

    wb = openpyxl.load_workbook(template_stream())
    ws = wb[SHEET_NAME]

    # ── Metadata block (Startup Report tab only) ────────────────────────────────
    for field, ref in META_CELLS.items():
        _set(ws, ref, getattr(meta, field, "") or "")

    # ── Valve rows: columns B..K from row 15 down (A is auto-numbered) ──────────
    for i, rec in enumerate(records):
        r = FIRST_DATA_ROW + i
        _set(ws, f"B{r}", getattr(rec, "valve_tag", "") or "")
        _set(ws, f"C{r}", map_product_line(getattr(rec, "valve_type", "") or ""))
        _set(ws, f"D{r}", getattr(rec, "model", "") or "")
        _set(ws, f"E{r}", map_valve_type(getattr(rec, "valve_type", "") or ""))
        _set(ws, f"F{r}", getattr(rec, "location_room", "") or "")
        _set(ws, f"G{r}", map_pass_fail(getattr(rec, "pass_fail", "") or ""))
        _set(ws, f"H{r}", _num_or_text(getattr(rec, "valve_min_sp", "")))
        _set(ws, f"I{r}", _num_or_text(getattr(rec, "valve_max_sp", "")))
        # J (Face Velocity FPM) — blank for v1
        _set(ws, f"K{r}", getattr(rec, "notes", "") or "")

    # Notes column (K) wraps long text on the data rows (preserve other alignment).
    _enable_notes_wrap(ws)

    wb.save(output_path)
