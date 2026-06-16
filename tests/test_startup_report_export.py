"""
Tests for the Startup Report export (startup_report_export.py).

Qt-free: exercises the headless export engine + the embedded template. Uses stdlib
unittest (pytest is not a project dependency). Run with:

    .venv\\Scripts\\python.exe -m unittest tests.test_startup_report_export -v
"""

from __future__ import annotations

import os
import sys
import tempfile
import unittest
from types import SimpleNamespace

# Make the repo root importable when run from anywhere.
_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

import openpyxl

import startup_report_export as sr
from startup_report_export import (
    StartupReportMeta, TooManyValvesError, MAX_VALVES,
    export_startup_report, prefill_meta, derive_product_lines,
    map_product_line, map_valve_type, map_pass_fail, generate_executive_summary,
)
from startup_report_template import template_stream


def make_record(**kw):
    """A duck-typed checkout record with all attributes the engine reads."""
    base = dict(valve_tag="", project="", ats_job_number="", date="", technician="",
                description="", location_room="", model="", valve_type="Fume Hood",
                pass_fail="", valve_min_sp="", valve_max_sp="", notes="", job_id="J1")
    base.update(kw)
    return SimpleNamespace(**base)


def export_to_temp(meta, records):
    """Export and return a reloaded (output_wb, startup_ws)."""
    tmpdir = tempfile.mkdtemp(prefix="sr_test_")
    out = os.path.join(tmpdir, "out.xlsx")
    export_startup_report(meta, records, out)
    wb = openpyxl.load_workbook(out)
    return wb, wb[sr.SHEET_NAME], out


class TemplateEmbeddingTests(unittest.TestCase):
    def test_template_opens_from_embedded(self):
        wb = openpyxl.load_workbook(template_stream())
        self.assertEqual(wb.sheetnames, ["Cover", "Startup Report"])


class MappingTests(unittest.TestCase):
    def test_pass_fail_mapping(self):
        self.assertEqual(map_pass_fail("Pass"), "PASS")
        self.assertEqual(map_pass_fail("Fail"), "FAIL")
        self.assertEqual(map_pass_fail(""), "")
        self.assertEqual(map_pass_fail("N/A"), "")

    def test_valve_type_mapping_all_nine(self):
        self.assertEqual(map_valve_type("Fume Hood"), "Fume Hood")
        self.assertEqual(map_valve_type("CSCP Fume Hood"), "Fume Hood")
        self.assertEqual(map_valve_type("PBC Room"), "PBC")
        self.assertEqual(map_valve_type("MAV"), "Supply")
        self.assertEqual(map_valve_type("GEX"), "General Exhaust")
        self.assertEqual(map_valve_type("Snorkel"), "General Exhaust")
        self.assertEqual(map_valve_type("Canopy"), "General Exhaust")
        self.assertEqual(map_valve_type("Draw Down Bench"), "General Exhaust")
        self.assertEqual(map_valve_type("Gas Cabinet"), "General Exhaust")
        self.assertEqual(map_valve_type("Unknown Type"), "")

    def test_valve_type_values_are_in_dropdown(self):
        allowed = {"Supply", "General Exhaust", "Fume Hood", "PBC", ""}
        for vt in ["Fume Hood", "CSCP Fume Hood", "PBC Room", "MAV", "GEX",
                   "Snorkel", "Canopy", "Draw Down Bench", "Gas Cabinet", "Whatever"]:
            self.assertIn(map_valve_type(vt), allowed)

    def test_product_line_mapping(self):
        self.assertEqual(map_product_line("CSCP Fume Hood"), "CSCP")
        self.assertEqual(map_product_line("PBC Room"), "CSCP")
        self.assertEqual(map_product_line("Fume Hood"), "Celeris")
        self.assertEqual(map_product_line("GEX"), "Celeris")
        self.assertEqual(map_product_line("Unknown"), "")

    def test_derive_product_lines(self):
        self.assertEqual(derive_product_lines([make_record(valve_type="Fume Hood")]), "Celeris")
        self.assertEqual(derive_product_lines([make_record(valve_type="PBC Room")]), "CSCP")
        self.assertEqual(
            derive_product_lines([make_record(valve_type="Fume Hood"),
                                  make_record(valve_type="CSCP Fume Hood")]),
            "CSCP & Celeris",
        )
        self.assertEqual(derive_product_lines([]), "")


class MetadataWriteTests(unittest.TestCase):
    def test_metadata_cells_populated(self):
        meta = StartupReportMeta(
            project="Memorial Lab", site_name="Main Campus", building="B",
            floor="3", technician="J. Glave", ats_job_number="ATS-2026-0142",
            date_of_checkout="2026-06-16", product_lines="Celeris",
            description="Lab wing", executive_summary="All passed.",
        )
        wb, ws, _ = export_to_temp(meta, [])
        self.assertEqual(ws["C3"].value, "Memorial Lab")
        self.assertEqual(ws["C4"].value, "Main Campus")
        self.assertEqual(ws["C5"].value, "B")
        self.assertEqual(ws["C6"].value, "3")
        self.assertEqual(ws["C7"].value, "J. Glave")
        self.assertEqual(ws["C8"].value, "ATS-2026-0142")
        self.assertEqual(ws["C9"].value, "2026-06-16")
        self.assertEqual(ws["C10"].value, "Celeris")
        self.assertEqual(ws["C11"].value, "Lab wing")
        self.assertEqual(ws["G4"].value, "All passed.")

    def test_missing_optional_fields_blank(self):
        # Only project filled; everything else blank should remain unwritten (None).
        wb, ws, _ = export_to_temp(StartupReportMeta(project="P"), [])
        self.assertEqual(ws["C3"].value, "P")
        for ref in ("C4", "C5", "C6", "C7", "C8", "C9", "C11", "G4"):
            self.assertIsNone(ws[ref].value, f"{ref} should be blank")


class ValveRowTests(unittest.TestCase):
    def test_valve_rows_b_to_k(self):
        rec = make_record(valve_tag="03-SV-001", valve_type="MAV", model="CELERIS 2",
                          pass_fail="Pass", valve_min_sp="120", valve_max_sp="900",
                          notes="ok")
        wb, ws, _ = export_to_temp(StartupReportMeta(), [rec])
        self.assertEqual(ws["B15"].value, "03-SV-001")
        self.assertEqual(ws["C15"].value, "Celeris")      # MAV -> Celeris product line
        self.assertEqual(ws["D15"].value, "CELERIS 2")
        self.assertEqual(ws["E15"].value, "Supply")       # MAV -> Supply
        self.assertIsNone(ws["F15"].value)                # Location blank v1
        self.assertEqual(ws["G15"].value, "PASS")
        self.assertEqual(ws["H15"].value, 120)            # numeric
        self.assertEqual(ws["I15"].value, 900)
        self.assertIsNone(ws["J15"].value)                # Face Velocity blank v1
        self.assertEqual(ws["K15"].value, "ok")

    def test_column_a_never_written(self):
        rec = make_record(valve_tag="T1")
        wb, ws, _ = export_to_temp(StartupReportMeta(), [rec])
        # A keeps the template's auto-number formula, not a value we wrote.
        self.assertEqual(ws["A15"].value, '=IF(B15="","",ROW()-14)')

    def test_unset_pass_fail_blank(self):
        rec = make_record(valve_tag="T1", pass_fail="")
        wb, ws, _ = export_to_temp(StartupReportMeta(), [rec])
        self.assertIsNone(ws["G15"].value)

    def test_blank_min_max_not_written(self):
        rec = make_record(valve_tag="T1", valve_min_sp="", valve_max_sp="")
        wb, ws, _ = export_to_temp(StartupReportMeta(), [rec])
        self.assertIsNone(ws["H15"].value)
        self.assertIsNone(ws["I15"].value)

    def test_face_velocity_never_inferred(self):
        # Even if a record carried verification data, J stays blank in v1.
        rec = make_record(valve_tag="T1")
        rec.verification = {"face_velocity_result": "Pass", "face_velocity_notes": "95 fpm"}
        wb, ws, _ = export_to_temp(StartupReportMeta(), [rec])
        self.assertIsNone(ws["J15"].value)

    def test_zero_valves(self):
        wb, ws, _ = export_to_temp(StartupReportMeta(project="P"), [])
        self.assertIsNone(ws["B15"].value)

    def test_one_valve(self):
        wb, ws, _ = export_to_temp(StartupReportMeta(), [make_record(valve_tag="ONE")])
        self.assertEqual(ws["B15"].value, "ONE")
        self.assertIsNone(ws["B16"].value)

    def test_forty_valves_fills_last_row(self):
        recs = [make_record(valve_tag=f"V{i:02d}", pass_fail="Pass") for i in range(MAX_VALVES)]
        wb, ws, _ = export_to_temp(StartupReportMeta(), recs)
        self.assertEqual(ws["B15"].value, "V00")
        self.assertEqual(ws["B54"].value, f"V{MAX_VALVES - 1:02d}")  # row 54 = last preformatted

    def test_forty_one_valves_blocked(self):
        recs = [make_record(valve_tag=f"V{i}") for i in range(MAX_VALVES + 1)]
        with self.assertRaises(TooManyValvesError):
            export_startup_report(StartupReportMeta(), recs,
                                  os.path.join(tempfile.mkdtemp(), "x.xlsx"))


class CoverAndPreservationTests(unittest.TestCase):
    def test_cover_tab_unchanged(self):
        # Capture Cover formulas from the pristine template.
        tpl = openpyxl.load_workbook(template_stream())["Cover"]
        before = {c: tpl[c].value for c in ("C6", "C7", "C12", "B16")}
        wb, _, _ = export_to_temp(
            StartupReportMeta(project="P", executive_summary="X"),
            [make_record(valve_tag="T1", pass_fail="Pass")],
        )
        cover = wb["Cover"]
        for c, v in before.items():
            self.assertEqual(cover[c].value, v, f"Cover {c} changed")
        # Cover values are still formulas, not literals we wrote.
        self.assertTrue(str(cover["C6"].value).startswith("="))

    def test_template_features_preserved(self):
        wb, ws, _ = export_to_temp(
            StartupReportMeta(project="P"),
            [make_record(valve_tag="T1", valve_type="Fume Hood", pass_fail="Fail")],
        )
        # Data validations (3 dropdowns) survive.
        dvs = ws.data_validations.dataValidation
        self.assertEqual(len(dvs), 3)
        sqrefs = {str(dv.sqref) for dv in dvs}
        self.assertEqual(sqrefs, {"C15:C54", "E15:E54", "G15:G54"})
        # Conditional formatting on G15:G54 survives.
        cf_ranges = [str(r) for r in ws.conditional_formatting]
        self.assertTrue(any("G15:G54" in r for r in cf_ranges))
        # Key merge, hidden detail group, print area, repeat header.
        self.assertIn("G4:K11", [str(m) for m in ws.merged_cells.ranges])
        self.assertTrue(ws.column_dimensions["L"].hidden)
        self.assertEqual(ws.column_dimensions["L"].outlineLevel, 1)
        self.assertTrue(ws.column_dimensions["AE"].hidden)
        self.assertEqual(ws.print_area, "'Startup Report'!$A$1:$K$56")
        self.assertEqual(ws.print_title_rows, "$14:$14")


class RealRecordCompatibilityTests(unittest.TestCase):
    def test_real_valvecheckout_exports(self):
        # Confirms the engine's attribute names match the real dataclass.
        from checkout_tool_backend import ValveCheckout
        rec = ValveCheckout(valve_tag="REAL-1", valve_type="GEX", model="CELERIS 2",
                            pass_fail="Pass", valve_min_sp="50", valve_max_sp="500",
                            notes="real record")
        meta = prefill_meta(rec, [rec])
        self.assertEqual(meta.product_lines, "Celeris")
        wb, ws, _ = export_to_temp(meta, [rec])
        self.assertEqual(ws["B15"].value, "REAL-1")
        self.assertEqual(ws["E15"].value, "General Exhaust")  # GEX -> General Exhaust
        self.assertEqual(ws["G15"].value, "PASS")


class LocationRoomTests(unittest.TestCase):
    def test_field_exists_with_blank_default(self):
        from checkout_tool_backend import ValveCheckout
        self.assertIn("location_room", ValveCheckout.__dataclass_fields__)
        self.assertEqual(ValveCheckout().location_room, "")

    def test_old_record_without_location_room_loads(self):
        # Mirrors CheckoutStore._load: filter unknown keys by __dataclass_fields__.
        from checkout_tool_backend import ValveCheckout
        old = {"id": "x", "valve_tag": "V1", "valve_type": "Fume Hood"}  # no location_room
        rec = ValveCheckout(**{k: v for k, v in old.items()
                               if k in ValveCheckout.__dataclass_fields__})
        self.assertEqual(rec.location_room, "")
        self.assertEqual(rec.valve_tag, "V1")

    def test_location_room_round_trips_on_real_record(self):
        from checkout_tool_backend import ValveCheckout
        from dataclasses import asdict
        rec = ValveCheckout(valve_tag="V1", location_room="Lab 314")
        reloaded = ValveCheckout(**asdict(rec))
        self.assertEqual(reloaded.location_room, "Lab 314")

    def test_location_room_writes_to_column_F(self):
        rec = make_record(valve_tag="V1", location_room="Lab 314")
        wb, ws, _ = export_to_temp(StartupReportMeta(), [rec])
        self.assertEqual(ws["F15"].value, "Lab 314")

    def test_blank_location_room_not_written(self):
        rec = make_record(valve_tag="V1", location_room="")
        wb, ws, _ = export_to_temp(StartupReportMeta(), [rec])
        self.assertIsNone(ws["F15"].value)


class RemovedFieldsTests(unittest.TestCase):
    # Operator-flagged columns that must NOT be written by the app.
    REMOVED = {
        "M": "Mag Diff Pressure", "R": "Primary Air from AHU", "T": "Htg Offset",
        "U": "Clg Offset", "X": "Exhaust Damper Pos.", "Z": "Verify Schedule",
        "Y": "Air Differential", "AC": "Linkage Cotter Pins",
    }

    def test_removed_fields_not_written(self):
        rec = make_record(valve_tag="V1", valve_type="Fume Hood", pass_fail="Pass",
                          notes="some notes", location_room="Rm 1")
        wb, ws, _ = export_to_temp(StartupReportMeta(project="P"), [rec])
        for col, name in self.REMOVED.items():
            self.assertIsNone(ws[f"{col}15"].value, f"{col} ({name}) should be blank")

    def test_no_hidden_detail_column_written(self):
        # The whole L..AE band stays blank on a populated data row.
        from openpyxl.utils import get_column_letter
        rec = make_record(valve_tag="V1", pass_fail="Pass", notes="n", location_room="r")
        wb, ws, _ = export_to_temp(StartupReportMeta(), [rec])
        for c in range(12, 32):  # L (12) .. AE (31)
            self.assertIsNone(ws[f"{get_column_letter(c)}15"].value,
                              f"col {get_column_letter(c)} should be blank")


class NotesWrapTests(unittest.TestCase):
    def test_notes_column_wraps_all_data_rows(self):
        rec = make_record(valve_tag="V1", notes="a long note\nsecond line")
        wb, ws, _ = export_to_temp(StartupReportMeta(), [rec])
        for r in (15, 16, 30, 54):
            self.assertTrue(ws[f"K{r}"].alignment.wrap_text, f"K{r} should wrap")

    def test_wrap_does_not_break_other_columns(self):
        rec = make_record(valve_tag="V1", notes="x")
        wb, ws, _ = export_to_temp(StartupReportMeta(), [rec])
        # B column shouldn't have been forced to wrap by the Notes-only logic.
        self.assertEqual(ws["B15"].value, "V1")


class ExecutiveSummaryTests(unittest.TestCase):
    def test_writes_to_g4(self):
        wb, ws, _ = export_to_temp(StartupReportMeta(executive_summary="Hello"), [])
        self.assertEqual(ws["G4"].value, "Hello")

    def test_summary_with_failures_lists_tags(self):
        recs = [make_record(valve_tag="VAV-101", pass_fail="Pass"),
                make_record(valve_tag="VAV-103", pass_fail="Fail"),
                make_record(valve_tag="FH-204", pass_fail="Fail")]
        s = generate_executive_summary(recs)
        self.assertIn("3 valves checked", s)
        self.assertIn("1 passed, 2 failed", s)
        self.assertIn("VAV-103", s)
        self.assertIn("FH-204", s)
        self.assertIn("See Notes column", s)

    def test_summary_all_pass(self):
        recs = [make_record(valve_tag="A", pass_fail="Pass"),
                make_record(valve_tag="B", pass_fail="Pass")]
        self.assertEqual(generate_executive_summary(recs),
                         "All checked valves passed. No issues noted.")

    def test_summary_empty(self):
        self.assertEqual(generate_executive_summary([]), "No valves checked.")

    def test_summary_unset_no_failures(self):
        recs = [make_record(valve_tag="A", pass_fail="Pass"),
                make_record(valve_tag="B", pass_fail="")]
        s = generate_executive_summary(recs)
        self.assertIn("0 failed", s)
        self.assertIn("No issues noted", s)

    def test_summary_does_not_invent_issues(self):
        # No "Fail" anywhere -> never mentions issues/failures count > 0.
        recs = [make_record(valve_tag="A", pass_fail="Pass")]
        s = generate_executive_summary(recs)
        self.assertNotIn("failed.", s.replace("0 failed.", ""))
        self.assertNotIn("Issue", s)

    def test_prefill_includes_generated_summary(self):
        recs = [make_record(valve_tag="A", pass_fail="Pass"),
                make_record(valve_tag="BAD-1", pass_fail="Fail")]
        meta = prefill_meta(recs[0], recs)
        self.assertIn("BAD-1", meta.executive_summary)
        self.assertIn("1 failed", meta.executive_summary)


if __name__ == "__main__":
    unittest.main(verbosity=2)
