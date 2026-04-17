"""
checkout_export.py — Export ValveCheckout records to Excel.

Loads checkout_template.xlsx (bundled with the app), fills in data from one
or more ValveCheckout records (one sheet per record), and saves to a new file.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from checkout_tool_backend import ValveCheckout


# ── Template location ──────────────────────────────────────────────────────────

TEMPLATE_NAME = "checkout_template.xlsx"

_VALVE_TYPE_TEMPLATE: dict[str, str] = {
    "Fume Hood":        "checkout_template.xlsx",
    "GEX":              "template_gex.xlsx",
    "MAV":              "template_mav.xlsx",
    "Snorkel":          "checkout_template.xlsx",
    "Canopy":           "checkout_template.xlsx",
    "Draw Down Bench":  "checkout_template.xlsx",
    "Gas Cabinet":      "checkout_template.xlsx",
    "CSCP Fume Hood":   "template_cscp_fh.xlsx",
    "PBC Room":         "template_pbc_room.xlsx",
}

CHECK = "\u2714"   # ✔  (same character used in the template's factory rows)


def _resource_path(filename: str) -> str:
    if getattr(sys, "frozen", False):
        base = Path(getattr(sys, "_MEIPASS", ""))
    else:
        base = Path(__file__).parent
    return str(base / filename)


# ── Wiring row maps ────────────────────────────────────────────────────────────
# Maps the wiring-list index (same index stored in record.wiring keys)
# to the Excel row number where that terminal row lives in the template.

# Phoenix Air Valve (columns A–F; Install=col D=4, Wired=col E=5)
_PHOENIX_ROW: dict[int, int] = {
    # TB1 — INPUTS  (rows 11-18)
    0: 11, 1: 12, 2: 13, 3: 14, 4: 15, 5: 16, 6: 17, 7: 18,
    # TB2 — OUTPUTS  (rows 20-26)
    8: 20, 9: 21, 10: 22, 11: 23, 12: 24, 13: 25, 14: 26,
    # TB3 — INTERNAL factory rows (28-32) → no checkboxes written
    # TB4 — POWER  (rows 34-36)
    20: 34, 21: 35, 22: 36,
    # TB6 — COMM  (rows 38-39)
    23: 38, 24: 39,
    # TB7 — ACTUATOR factory rows (41-42) → no checkboxes written
}

# Black Box (columns G–L; Install=col J=10, Wired=col K=11)
_BB_ROW: dict[int, int] = {
    # TB1 — POWER IN  (rows 11-13)
    0: 11, 1: 12, 2: 13,
    # TB2 — OUTPUTS  (rows 15-16)
    3: 15, 4: 16,
    # TB3 — INPUTS  (rows 18-19)
    5: 18, 6: 19,
    # FHM Sentry — Fume Hood Monitor  (rows 21-24)
    7: 21, 8: 22, 9: 23, 10: 24,
}

# Configuration (key → row; CFM=col I=9, Notes=col J=10)
_CONFIG_ROW: dict[str, int] = {
    "valve_min":     30,
    "valve_max":     31,
    "sched_min":     32,
    "sched_max":     33,
    "hood_sash_min": 34,
    "hood_sash_max": 35,
}

# Verification (key → row; Result=col I=9, Notes=col J=10)
_VERIFY_ROW: dict[str, int] = {
    "face_velocity":      37,
    "sash_height_alarm":  38,
    "sash_sensor_output": 39,
    "low_flow_alarm":     40,
    "jam_alarm":          41,
    "emergency_exhaust":  42,
    "mute_function":      43,
}

# Notes rows (rows 45-49, 5 lines; each row is A:L merged)
_NOTE_ROWS = [45, 46, 47, 48, 49]


# ── Sheet filler ──────────────────────────────────────────────────────────────

def _w(ws: Worksheet, row: int, col: int, value) -> None:
    """Write value to cell without disturbing its existing formatting."""
    ws.cell(row=row, column=col).value = value


def _safe_tab_name(valve_tag: str) -> str:
    name = (valve_tag or "Checkout")[:31]
    for ch in r"\/?*[]":
        name = name.replace(ch, "_")
    return name or "Checkout"


def fill_sheet(ws: Worksheet, record: ValveCheckout) -> None:
    """Populate a Fume Hood / fallback worksheet (12-column template)."""

    # Header — data cells confirmed against checkout_template.xlsx merged ranges
    _w(ws, 2,  3, record.project)
    _w(ws, 2,  9, record.ats_job_number)
    _w(ws, 3,  3, record.valve_tag)
    _w(ws, 3,  9, record.date)
    _w(ws, 4,  3, record.technician)
    _w(ws, 4,  9, record.description)
    _w(ws, 5,  3, record.model or "CELERIS 2")
    # Row 7: Pass/Fail=A, Emer.Min=C, Valve Min=E, Valve Max=H
    _w(ws, 7,  1, record.pass_fail)
    _w(ws, 7,  3, record.emer_min)
    _w(ws, 7,  5, record.valve_min_sp)
    _w(ws, 7,  8, record.valve_max_sp)

    # Phoenix wiring — Install=col D(4), Wired=col E(5)
    w = record.wiring
    for idx, row in _PHOENIX_ROW.items():
        _w(ws, row, 4, CHECK if w.get(f"p_{idx}_i") else "")
        _w(ws, row, 5, CHECK if w.get(f"p_{idx}_w") else "")

    # Black Box wiring — Install=col J(10), Wired=col K(11)
    for idx, row in _BB_ROW.items():
        _w(ws, row, 10, CHECK if w.get(f"b_{idx}_i") else "")
        _w(ws, row, 11, CHECK if w.get(f"b_{idx}_w") else "")

    # Sash sensor — merged area G25:L26 is the data cell
    _w(ws, 25, 7, (CHECK + "  Mounting Complete") if record.sash_sensor_mounted else "")

    # Configuration — CFM=col I(9), Notes=col J(10)
    cfg = record.config
    for key, row in _CONFIG_ROW.items():
        _w(ws, row, 9,  cfg.get(f"{key}_cfm",   ""))
        _w(ws, row, 10, cfg.get(f"{key}_notes",  ""))

    # Verification — Result=col I(9), Notes=col J(10)
    vfy = record.verification
    for key, row in _VERIFY_ROW.items():
        _w(ws, row, 9,  vfy.get(f"{key}_result", ""))
        _w(ws, row, 10, vfy.get(f"{key}_notes",  ""))

    # Notes — rows 45-49, each row A:L merged; write to col A(1)
    lines = (record.notes or "").split("\n")
    for i, note_row in enumerate(_NOTE_ROWS):
        _w(ws, note_row, 1, lines[i] if i < len(lines) else "")

    ws.title = _safe_tab_name(record.valve_tag)


def fill_sheet_gex_mav(ws: Worksheet, record: ValveCheckout) -> None:
    """Populate a GEX or MAV worksheet (6-column template, Phoenix wiring only).

    Audited against Phoenix_Celeris_GEX_Checkout_Reformatted.xlsx and
    Phoenix_Celeris_MAV_Checkout_Reformatted.xlsx. Both share the same layout.

    Header merged ranges:
      Row 2: A2:B2 = "Project" label, C2 = project value,
             D2 = "ATS Job Number" label, E2:F2 = ATS job # value
      Row 3: A3:B3 = "Valve Tag #" label, C3 = valve tag value,
             D3 = "Date" label, E3:F3 = date value
      Row 4: A4:B4 = "Technician" label, C4 = technician value,
             D4 = "Description" label, E4:F4 = description value
      Row 5: A5:B5 = "Model" label, C5:F5 = model value
      Row 6: labels — A6=Pass/Fail, B6=Emer.Min, C6:D6=Valve Min, E6:F6=Valve Max
      Row 7: data  — A7=Pass/Fail, B7=Emer.Min, C7:D7=Valve Min, E7:F7=Valve Max
    Wiring: Phoenix only, Install=col D(4), Wired=col E(5), same rows as Fume Hood.
    Notes: rows 45-49, each A:F merged; write to col A(1).
    No Black Box, no Config section, no Verification section.
    """

    # Header
    _w(ws, 2,  3, record.project)
    _w(ws, 2,  5, record.ats_job_number)
    _w(ws, 3,  3, record.valve_tag)
    _w(ws, 3,  5, record.date)
    _w(ws, 4,  3, record.technician)
    _w(ws, 4,  5, record.description)
    _w(ws, 5,  3, record.model or "CELERIS 2")
    # Row 7: Pass/Fail=A(1), Emer.Min=B(2), Valve Min=C(3), Valve Max=E(5)
    _w(ws, 7,  1, record.pass_fail)
    _w(ws, 7,  2, record.emer_min)
    _w(ws, 7,  3, record.valve_min_sp)
    _w(ws, 7,  5, record.valve_max_sp)

    # Phoenix wiring — Install=col D(4), Wired=col E(5)
    w = record.wiring
    for idx, row in _PHOENIX_ROW.items():
        _w(ws, row, 4, CHECK if w.get(f"p_{idx}_i") else "")
        _w(ws, row, 5, CHECK if w.get(f"p_{idx}_w") else "")

    # Notes — rows 45-49, each A:F merged; write to col A(1)
    lines = (record.notes or "").split("\n")
    for i, note_row in enumerate(_NOTE_ROWS):
        _w(ws, note_row, 1, lines[i] if i < len(lines) else "")

    ws.title = _safe_tab_name(record.valve_tag)


# ── CSCP Fume Hood row maps ────────────────────────────────────────────────────
# Audited against ACM_Fume_Hood_Reformatted.xlsx merged-cell ranges.
#
# ACM wiring — single "Wiring" checkbox at col F (6).
# Factory rows (P20: 14-16, P30: 17-21, P80: 39-41) have no checkboxes.
_ACM_ROW: dict[int, int] = {
    # Connector/Terminal P10  (rows 11-13)
    0: 11, 1: 12, 2: 13,
    # P20 (i=3,4) and P30 (i=5-8) are factory-wired — no entry
    # Connector/Terminal P40  (rows 23-25)
    9: 23, 10: 24, 11: 25,
    # Connector/Terminal P50  (rows 27-29)
    12: 27, 13: 28, 14: 29,
    # Connector/Terminal P60  (rows 31-33)
    15: 31, 16: 32, 17: 33,
    # Connector/Terminal P70  (rows 35-38)
    18: 35, 19: 36, 20: 37, 21: 38,
    # P80 (i=22-24) are factory-wired — no entry
}

# DHV Black Box — single "Wiring" checkbox at col L (12).
_DHV_ROW: dict[int, int] = {
    # TB1 POWER IN  (rows 11-13)
    0: 11, 1: 12, 2: 13,
    # TB2 OUTPUTS  (rows 15-16)
    3: 15, 4: 16,
    # TB3 INPUTS  (rows 18-22)
    5: 18, 6: 19, 7: 20, 8: 21, 9: 22,
    # UIO 1  (rows 24-25)
    10: 24, 11: 25,
    # UIO 2  (rows 27-28)
    12: 27, 13: 28,
    # UI 1  (rows 30-31)
    14: 30, 15: 31,
    # UI 2  (rows 33-34)
    16: 33, 17: 34,
    # DO SWITCH  (rows 36-37)
    18: 36, 19: 37,
    # MSTP  (rows 39-40)
    20: 39, 21: 40,
}

# Configuration  (key → row;  CFM = col E=5,  Notes = col I=9)
_CSCP_CONFIG_ROW: dict[str, int] = {
    "valve_min":     46,
    "valve_max":     47,
    "sched_min":     48,
    "sched_max":     49,
    "hood_sash_min": 50,
    "hood_sash_max": 51,
}

# Verification  (key → row;  Result = col E=5,  Notes = col I=9)
_CSCP_VERIFY_ROW: dict[str, int] = {
    "face_velocity":      53,
    "sash_height_alarm":  54,
    "sash_sensor_output": 55,
    "low_flow_alarm":     56,
    "jam_alarm":          57,
    "emergency_exhaust":  58,
}

# Notes rows 60-64 (5 lines, each A:L merged); write to col A (1)
_CSCP_NOTE_ROWS = [60, 61, 62, 63, 64]


def fill_sheet_cscp_fh(ws: Worksheet, record: ValveCheckout) -> None:
    """Populate a CSCP Fume Hood worksheet (ACM_Fume_Hood_Reformatted template).

    Header layout is identical to the Celeris Fume Hood template (rows 2-7).
    DHV Black Box wiring uses a single Wiring checkbox at col L (12).
    Config:  CFM = col E (5),  Notes = col I (9),  rows 46-51.
    Verify:  Result = col E (5),  Notes = col I (9),  rows 53-58.
    Sash sensor mounting check written to row 42, col G (7) on DHV side.
    Notes: rows 60-64, col A (1).
    """

    # Header — identical positions to Celeris Fume Hood template
    _w(ws, 2,  3, record.project)
    _w(ws, 2,  9, record.ats_job_number)
    _w(ws, 3,  3, record.valve_tag)
    _w(ws, 3,  9, record.date)
    _w(ws, 4,  3, record.technician)
    _w(ws, 4,  9, record.description)
    _w(ws, 5,  3, record.model or "ACM (CSCP)")
    # Row 7: Pass/Fail=A(1), Emer.Min=C(3), Valve Min=E(5), Valve Max=H(8)
    _w(ws, 7,  1, record.pass_fail)
    _w(ws, 7,  3, record.emer_min)
    _w(ws, 7,  5, record.valve_min_sp)
    _w(ws, 7,  8, record.valve_max_sp)

    # ACM wiring — single Wiring checkbox at col F (6)
    w = record.wiring
    for idx, row in _ACM_ROW.items():
        _w(ws, row, 6, CHECK if w.get(f"acm_{idx}_w") else "")

    # DHV Black Box wiring — single Wiring checkbox at col L (12)
    for idx, row in _DHV_ROW.items():
        _w(ws, row, 12, CHECK if w.get(f"dhv_{idx}_w") else "")

    # Sash sensor mounting — DHV side row 41, col G (7)
    _w(ws, 41, 7, (CHECK + "  Mounting Complete") if record.sash_sensor_mounted else "")

    # Configuration — CFM = col E (5), Notes = col I (9)
    cfg = record.config
    for key, row in _CSCP_CONFIG_ROW.items():
        _w(ws, row, 5, cfg.get(f"{key}_cfm",   ""))
        _w(ws, row, 9, cfg.get(f"{key}_notes",  ""))

    # Verification — Result = col E (5), Notes = col I (9)
    vfy = record.verification
    for key, row in _CSCP_VERIFY_ROW.items():
        _w(ws, row, 5, vfy.get(f"{key}_result", ""))
        _w(ws, row, 9, vfy.get(f"{key}_notes",  ""))

    # Notes — rows 60-64, col A (1)
    lines = (record.notes or "").split("\n")
    for i, note_row in enumerate(_CSCP_NOTE_ROWS):
        _w(ws, note_row, 1, lines[i] if i < len(lines) else "")

    ws.title = _safe_tab_name(record.valve_tag)


# ── PBC Room row maps ─────────────────────────────────────────────────────────
# Audited against PBC_Room_Checkout_Reformatted.xlsx.
#
# Header layout matches all other templates (rows 2-7, same cols).
# Row 7 has ONLY Pass/Fail at A7 — no Emer.Min, Valve Min, or Valve Max.
# Left panel: Install=col D(4), Wired=col E(5).
# Right panel: Install=col J(10), Wired=col K(11).
# Notes: rows 54-58, col A (1).  No Config or Verification section.

_PBC_L_ROW: dict[int, int] = {
    # TB1 24V Power  (rows 11-13)
    0: 11, 1: 12, 2: 13,
    # DO Relay NC1-IN4  (rows 15-26; row 14 is empty)
    3: 15, 4: 16, 5: 17, 6: 18, 7: 19, 8: 20,
    9: 21, 10: 22, 11: 23, 12: 24, 13: 25, 14: 26,
    # DO SSR  (rows 28-35; row 27 empty; i=16 factory=SRIN row 29)
    15: 28, 17: 30, 18: 31, 19: 32, 20: 33, 21: 34, 22: 35,
    # I/O ???  (rows 37-44; row 36 empty)
    23: 37, 24: 38, 25: 39, 26: 40, 27: 41, 28: 42, 29: 43, 30: 44,
    # UIO  (rows 46-52; row 45 empty)
    31: 46, 32: 47, 33: 48, 34: 49, 35: 50, 36: 51, 37: 52,
}

_PBC_R_ROW: dict[int, int] = {
    # TB1 POWER IN ???  (rows 11-13)
    0: 11, 1: 12, 2: 13,
    # BACnet MS/TP  (rows 15-17; row 14 empty)
    3: 15, 4: 16, 5: 17,
    # RS485  (rows 19-21; row 18 empty)
    6: 19, 7: 20, 8: 21,
    # Power 24  (rows 23-24; row 22 empty)
    9: 23, 10: 24,
    # SYLK  (rows 26-27; row 25 empty)
    11: 26, 12: 27,
    # UIO IO1-IO12 + 24VDC OUT  (rows 29-49; row 28 empty)
    13: 29, 14: 30, 15: 31, 16: 32, 17: 33, 18: 34,
    19: 35, 20: 36, 21: 37, 22: 38, 23: 39, 24: 40,
    25: 41, 26: 42, 27: 43, 28: 44, 29: 45, 30: 46,
    31: 47, 32: 48, 33: 49,
}

_PBC_NOTE_ROWS = [54, 55, 56, 57, 58]


def fill_sheet_pbc_room(ws: Worksheet, record: ValveCheckout) -> None:
    """Populate a PBC Room worksheet (PBC_Room_Checkout_Reformatted template).

    Header rows 2-5 are identical to all other templates.
    Row 7: Pass/Fail = A (1) only — no valve SP fields.
    Left wiring: Install=col D(4), Wired=col E(5).
    Right wiring: Install=col J(10), Wired=col K(11).
    Notes: rows 54-58, col A (1).
    """

    # Header
    _w(ws, 2,  3, record.project)
    _w(ws, 2,  9, record.ats_job_number)
    _w(ws, 3,  3, record.valve_tag)
    _w(ws, 3,  9, record.date)
    _w(ws, 4,  3, record.technician)
    _w(ws, 4,  9, record.description)
    _w(ws, 5,  3, record.model or "PBC (CSCP)")
    _w(ws, 7,  1, record.pass_fail)

    # PBC wiring — left panel: Install=col D(4), Wired=col E(5)
    w = record.wiring
    for idx, row in _PBC_L_ROW.items():
        _w(ws, row, 4, CHECK if w.get(f"pbc_l_{idx}_i") else "")
        _w(ws, row, 5, CHECK if w.get(f"pbc_l_{idx}_w") else "")

    # PBC wiring — right panel: Install=col J(10), Wired=col K(11)
    for idx, row in _PBC_R_ROW.items():
        _w(ws, row, 10, CHECK if w.get(f"pbc_r_{idx}_i") else "")
        _w(ws, row, 11, CHECK if w.get(f"pbc_r_{idx}_w") else "")

    # Notes — rows 54-58, col A (1)
    lines = (record.notes or "").split("\n")
    for i, note_row in enumerate(_PBC_NOTE_ROWS):
        _w(ws, note_row, 1, lines[i] if i < len(lines) else "")

    ws.title = _safe_tab_name(record.valve_tag)


# Types that use the GEX/MAV fill function
_GEX_MAV_TYPES = {"GEX", "MAV"}


# ── Public export entry point ─────────────────────────────────────────────────

def export_records(records: list[ValveCheckout], output_path: str) -> None:
    """
    Export one or more records to output_path (.xlsx).
    Produces one worksheet per record, each formatted like its valve-type template.
    """
    if not records:
        raise ValueError("No records provided for export.")

    template_name = _VALVE_TYPE_TEMPLATE.get(
        records[0].valve_type, TEMPLATE_NAME
    )
    template_path = _resource_path(template_name)
    wb = load_workbook(template_path)
    template_ws = wb.active

    sheets = [template_ws]
    for _ in records[1:]:
        sheets.append(wb.copy_worksheet(template_ws))

    for ws, record in zip(sheets, records):
        if record.valve_type in _GEX_MAV_TYPES:
            fill_sheet_gex_mav(ws, record)
        elif record.valve_type == "CSCP Fume Hood":
            fill_sheet_cscp_fh(ws, record)
        elif record.valve_type == "PBC Room":
            fill_sheet_pbc_room(ws, record)
        else:
            fill_sheet(ws, record)

    wb.save(output_path)
